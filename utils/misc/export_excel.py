from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import datetime
from loader import db
import re

def format_phone_number(phone_number):
    # Remove unwanted characters (spaces, parentheses, dashes)
    phone_number = re.sub(r'[^\d]', '', phone_number)
    
    # If the number starts with 998 (local format), add the + sign
    if phone_number.startswith("998"):
        phone_number = "+" + phone_number
    
    # If the number already starts with +998, no change is needed
    elif phone_number.startswith("8"):  # If the number starts with 8 (local without country code)
        phone_number = "+998" + phone_number[1:]
    
    # Mask the last 3 digits with asterisks
    if len(phone_number) > 4:  # Ensure there are enough digits to mask
        phone_number = phone_number[:-3] + "***"

    return phone_number
    
async def export_votes_to_excel():
    nominations_sql = """
    SELECT id, title
    FROM Nominations
    """
    nominations_data = await db.execute(nominations_sql, fetch=True)

    if not nominations_data:
        print("Nominatsiyalar mavjud emas.")
        return

    # Workbook yaratish
    wb = Workbook()
    wb.remove(wb.active)  # Dastlabki bitta sheetni o'chirish

    # Har bir nominatciya uchun alohida sheet yaratish
    for nomination in nominations_data:
        nomination_id = nomination['id']
        nomination_title = nomination['title']

        # Nominatsiya bo'yicha ovoz bergan foydalanuvchilarni olish
        sql = """
        SELECT u.full_name, u.telegram_id, u.phone, v.created_at
        FROM Votes v
        JOIN Users u ON v.user_id = u.id
        WHERE v.nomination_id = $1
        ORDER BY v.created_at DESC
        """
        users_data = await db.execute(sql, nomination_id, fetch=True)

        if not users_data:
            continue

        # Yangi sheet yaratish
        ws = wb.create_sheet(title=nomination_title[:31])  # Title limit to 31 chars

        # Parent columnni kengaytirish
        ws.merge_cells('A1:E1')
        parent_cell = ws['A1']
        parent_cell.value = f"Nomination: {nomination_title}"
        parent_cell.font = Font(bold=True)
        parent_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ikkinchi qatorda ustunlar nomini qo'shish
        columns = ["№", "Tolıq atı", "Telegram ID", "Telefon nomeri", "Dawıs bergen waqtı"]

        # Ustun sarlavhalarini bold qilish
        bold_font = Font(bold=True)
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_num, value=column_title)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Chegaralar va uslublar uchun Border
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        # Ma'lumotlarni markazlashtirish uchun Alignment
        alignment = Alignment(horizontal='center', vertical='center')

        # Foydalanuvchilarni qo'shish
        for idx, user in enumerate(users_data, start=1):
            formatted_phone = format_phone_number(user['phone'])
            # Agar created_at datetime tipida bo'lsa
            voice_time = user['created_at'].strftime('%d.%m.%Y %H:%M') if isinstance(user['created_at'], datetime.datetime) else user['created_at']

            row = [idx, user['full_name'], user['telegram_id'], formatted_phone, voice_time]
            ws.append(row)

            # Har bir hujayraga chegaralar va markazlashtirish qo'shish
            for cell in ws[idx + 2]:
                cell.border = border
                cell.alignment = alignment

        # Chegaralarni barcha ustunlar uchun qo'shish, shu jumladan ustun sarlavhalari
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

        # Ustunlarni moslashtirish
        for col in range(1, len(columns) + 1):
            max_length = 0
            column = get_column_letter(col)

            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)  # 2 — bir oz bo'shliq qo'shish
            ws.column_dimensions[column].width = adjusted_width

    # Excel faylga saqlash
    file_name = "storage/all_nominations_votes.xlsx"
    wb.save(file_name)

    print(f"{file_name} faylga saqlandi.")
    return file_name
